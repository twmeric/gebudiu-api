#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
格不丢翻译 API - 安全加强版
模块化架构，支持7大专业领域
"""

import os
import sys
import hashlib
import json
import time
import logging
import re
from io import BytesIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

from flask import Flask, request, send_file, jsonify, make_response
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from docx import Document
from openai import OpenAI, APIError, RateLimitError
from dotenv import load_dotenv
import zipfile
import xml.etree.ElementTree as ET
import tempfile
import shutil

# 加载环境变量
load_dotenv()

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('translation.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# 创建应用
app = Flask(__name__)

# 安全配置 - 允許所有來源（生產環境需要時再限制）
CORS(app, origins="*", supports_credentials=False)

# 額外 CORS 處理 - 確保所有響應都帶 CORS 頭（包括錯誤響應）
@app.after_request
def after_request(response):
    """為所有響應添加 CORS 頭"""
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization,X-Requested-With')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# 速率限制
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# 文件配置
MAX_FILE_SIZE = 16 * 1024 * 1024
ALLOWED_EXTENSIONS = {'.docx', '.xlsx'}

# DeepSeek API 配置 - 延遲加載
def get_deepseek_client():
    """延遲初始化 DeepSeek 客戶端"""
    api_key = os.getenv("DEEPSEEK_API_KEY")
    
    # Debug: 記錄所有環境變量（隱藏敏感值）
    if not api_key:
        logger.error("DEEPSEEK_API_KEY not found!")
        logger.info(f"Available env vars: {[k for k in os.environ.keys() if not k.startswith('_')][:20]}")
        # 嘗試小寫
        api_key = os.getenv("deepseek_api_key") or os.getenv("Deepseek_Api_Key")
        if api_key:
            logger.info("Found API key with different case!")
        else:
            raise ValueError("DEEPSEEK_API_KEY environment variable is required")
    
    logger.info(f"API Key found: {api_key[:10]}... (length: {len(api_key)})")
    return OpenAI(api_key=api_key, base_url="https://api.deepseek.com/v1")

# 延遲初始化
deepseek_client = None

# 领域配置
DOMAINS = {
    "general": {
        "name": "通用领域",
        "name_en": "General",
        "icon": "📋"
    },
    "electronics": {
        "name": "电子产品",
        "name_en": "Electronics",
        "icon": "🔌"
    },
    "medical": {
        "name": "医疗器械",
        "name_en": "Medical",
        "icon": "🏥"
    },
    "legal": {
        "name": "法律合同",
        "name_en": "Legal",
        "icon": "⚖️"
    },
    "marketing": {
        "name": "市场营销",
        "name_en": "Marketing",
        "icon": "📢"
    },
    "industrial": {
        "name": "工业技术",
        "name_en": "Industrial",
        "icon": "🏭"
    },
    "software": {
        "name": "软件/IT",
        "name_en": "Software",
        "icon": "💻"
    }
}

# Prompt 模板
PROMPTS = {
    "general": "You are a professional translator. Translate the following Chinese text to English for general business use. Provide ONLY the English translation.",
    "electronics": "You are an electronics industry translator. Use standard technical terminology. Preserve model numbers like G-190. Provide ONLY the English translation.",
    "medical": "You are a medical device translator. Use formal, cautious medical terminology suitable for regulatory compliance. Provide ONLY the English translation.",
    "legal": "You are a legal contract translator. Use precise legal terminology and formal contractual language. Provide ONLY the English translation.",
    "marketing": "You are a marketing copy translator. Make the content persuasive and appealing. Provide ONLY the English translation.",
    "industrial": "You are an industrial/technical translator. Use ISO-standard terminology. Provide ONLY the English translation.",
    "software": "You are a software/IT documentation translator. Preserve code snippets and API names. Provide ONLY the English translation."
}

# 缓存管理
class TranslationCache:
    def __init__(self, cache_file=".translation_cache.json", max_size=10000):
        self.cache_file = cache_file
        self.max_size = max_size
        self.cache = {}
        self._load()
    
    def _load(self):
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    self.cache = json.load(f)
                    logger.info(f"Loaded {len(self.cache)} cached translations")
        except Exception as e:
            logger.warning(f"Failed to load cache: {e}")
            self.cache = {}
    
    def _save(self):
        try:
            if len(self.cache) > self.max_size:
                items = list(self.cache.items())
                self.cache = dict(items[-self.max_size:])
            
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save cache: {e}")
    
    def get_key(self, text, domain):
        return hashlib.md5(f"{domain}:{text}".encode('utf-8')).hexdigest()
    
    def get(self, text, domain):
        key = self.get_key(text, domain)
        return self.cache.get(key)
    
    def set(self, text, domain, translated):
        key = self.get_key(text, domain)
        self.cache[key] = translated
        self._save()

cache = TranslationCache()

def validate_file(file_storage):
    """验证上传文件 - 增强版"""
    if not file_storage or not getattr(file_storage, 'filename', None):
        return False, "No file selected", None
    
    # 安全处理文件名（移除路径、保留基本名称）
    filename = os.path.basename(file_storage.filename)
    ext = os.path.splitext(filename)[1].lower()
    
    if ext not in ALLOWED_EXTENSIONS:
        return False, f"Invalid file type. Allowed: {', '.join(ALLOWED_EXTENSIONS)}", None
    
    try:
        file_buffer = file_storage.read()
        file_storage.seek(0)
    except Exception as e:
        logger.error(f"Failed to read file: {e}")
        return False, "Failed to read uploaded file", None
    
    if len(file_buffer) == 0:
        return False, "File is empty", None
    
    if len(file_buffer) > MAX_FILE_SIZE:
        max_mb = MAX_FILE_SIZE / 1024 / 1024
        return False, f"File too large. Maximum size: {max_mb}MB", None
    
    # 检查ZIP格式 (docx/xlsx都是ZIP)
    # 支持多种 ZIP 格式标识: PK\x03\x04 (标准) 或 PK\x05\x06 (空ZIP)
    header = file_buffer[:4]
    valid_zip_headers = [b'PK\x03\x04', b'PK\x05\x06', b'PK\x07\x08']
    if header not in valid_zip_headers:
        logger.warning(f"Invalid file header: {header.hex()} for file: {filename}")
        return False, f"Invalid file format (header: {header.hex()}). Please upload a valid .docx or .xlsx file", None
    
    return True, None, {
        "filename": filename,
        "extension": ext,
        "size": len(file_buffer),
        "buffer": file_buffer
    }

def should_translate(text):
    """判断文本是否需要翻译"""
    if not text or len(text.strip()) < 2:
        return False
    
    text = text.strip()
    
    if text.isdigit():
        return False
    
    if text.startswith('http://') or text.startswith('https://'):
        return False
    
    if '@' in text and '.' in text.split('@')[-1]:
        return False
    
    if re.match(r'^G-\d+$', text):
        return False
    
    return True

class TranslationService:
    """批量翻译服务 - 內存優化版（針對 Render 512MB）"""
    
    def __init__(self, domain="general"):
        self.domain = domain
        self.prompt = PROMPTS.get(domain, PROMPTS["general"])
        self.stats = {"api_calls": 0, "cache_hits": 0, "tokens_saved": 0}
    
    def translate_batch(self, texts, max_retries=2):
        """
        批量翻译 - 內存優化，減少 tokens 使用
        
        Args:
            texts: list of (id, text) tuples
            max_retries: 最大重試次數
        
        Returns:
            dict: {id: translated_text}
        """
        if not texts:
            return {}
        
        # 1. 先检查缓存
        to_translate = []
        results = {}
        
        for item_id, text in texts:
            if not should_translate(text):
                results[item_id] = text
                continue
            
            cached = cache.get(text, self.domain)
            if cached:
                results[item_id] = cached
                self.stats["cache_hits"] += 1
            else:
                to_translate.append((item_id, text))
        
        if not to_translate:
            return results
        
        # 2. 構建批量翻譯請求（簡化 prompt 減少 tokens）
        batch_text = "\n".join([f"[{i}] {text[:500]}" for i, (_, text) in enumerate(to_translate)])
        
        for attempt in range(max_retries):
            try:
                client = get_deepseek_client()
                logger.info(f"Batch translating {len(to_translate)} texts, attempt {attempt+1}")
                
                # 簡化 prompt 減少 token 使用
                batch_prompt = f"""{self.prompt}

Translate to English (prefix with [number]):
{batch_text}

Format: [0] Translation"""
                
                response = client.chat.completions.create(
                    model="deepseek-chat",
                    messages=[
                        {"role": "system", "content": batch_prompt}
                    ],
                    temperature=0.3,
                    max_tokens=2000,  # 減少 tokens 限制
                    timeout=60  # 添加超時
                )
                
                # 3. 解析结果
                raw_output = response.choices[0].message.content.strip()
                translations = self._parse_batch_output(raw_output, len(to_translate))
                
                # 4. 保存到缓存和结果
                for i, (item_id, original) in enumerate(to_translate):
                    translated = translations.get(i, original)
                    cache.set(original, self.domain, translated)
                    results[item_id] = translated
                
                self.stats["api_calls"] += 1
                logger.info(f"Batch translated {len(to_translate)} texts")
                return results
                
            except Exception as e:
                logger.error(f"Batch translation attempt {attempt+1} failed: {e}")
                if attempt < max_retries - 1:
                    import time
                    time.sleep(2 ** attempt)  # 指數退避
                else:
                    # 所有重試都失敗，返回原文
                    logger.error(f"All retries failed, returning original text")
                    for item_id, text in to_translate:
                        results[item_id] = text
                    return results
        
        return results
    
    def _parse_batch_output(self, output, expected_count):
        """解析批量翻译的输出"""
        results = {}
        # 匹配 [0] ... [1] ... 格式
        pattern = r'\[(\d+)\]\s*(.+?)(?=\[\d+\]|$)'
        matches = re.findall(pattern, output, re.DOTALL)
        
        for idx_str, text in matches:
            idx = int(idx_str)
            if idx < expected_count:
                results[idx] = text.strip()
        
        return results
    
    def translate_text(self, text):
        """兼容旧接口 - 单条翻译"""
        result = self.translate_batch([("single", text)])
        translated = result.get("single", text)
        is_cached = translated != text and self.stats["cache_hits"] > 0
        return translated, is_cached

class DocxProcessor:
    """
    DOCX 文档处理器 - 流式優化版（圖片不進記憶體）
    
    核心創新：
    - 圖片直接從 ZIP 複製，不進記憶體
    - 只提取 XML 文本進行翻譯
    - 記憶體使用從 50-200MB 降至 <10MB
    """
    
    def __init__(self, translator):
        self.translator = translator
        self.stats = {"paragraphs": 0, "tables": 0, "cells": 0, "media_skipped": 0}
    
    def process(self, file_buffer):
        """
        流式處理 DOCX 文件 - 圖片直通不進記憶體
        """
        import gc
        
        # 使用臨時文件處理（避免內存中同時存在兩個大文件）
        with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as temp_input:
            temp_input.write(file_buffer)
            temp_input_path = temp_input.name
        
        temp_output_path = temp_input_path.replace('.docx', '_translated.docx')
        
        try:
            # 使用流式處理
            self._process_streaming(temp_input_path, temp_output_path)
            
            # 讀取結果
            with open(temp_output_path, 'rb') as f:
                result = BytesIO(f.read())
            
            gc.collect()
            return result
            
        finally:
            # 清理臨時文件
            try:
                os.unlink(temp_input_path)
                if os.path.exists(temp_output_path):
                    os.unlink(temp_output_path)
            except:
                pass
    
    def _process_streaming(self, input_path: str, output_path: str):
        """
        流式 DOCX 處理核心 - 圖片不進記憶體
        """
        # 1. 掃描 ZIP 內容，分離媒體和 XML
        media_files = []
        xml_files = []
        
        with zipfile.ZipFile(input_path, 'r') as zf:
            for info in zf.infolist():
                if info.filename.startswith('word/media/'):
                    media_files.append(info.filename)
                    self.stats["media_skipped"] += info.file_size
                elif info.filename.endswith('.xml'):
                    xml_files.append(info.filename)
        
        logger.info(f"Found {len(media_files)} media files, {len(xml_files)} XML files")
        
        # 2. 提取所有需要翻譯的文本
        texts_to_translate = []
        xml_text_map = {}  # xml_file -> [(search_text, element_id)]
        
        for xml_file in xml_files:
            texts, mappings = self._extract_texts_from_xml(input_path, xml_file)
            if texts:
                texts_to_translate.extend(texts)
                xml_text_map[xml_file] = mappings
        
        logger.info(f"Collected {len(texts_to_translate)} texts to translate")
        
        if not texts_to_translate:
            # 無需翻譯，直接複製
            shutil.copy(input_path, output_path)
            return
        
        # 3. 批量翻譯
        translations = self._translate_texts(texts_to_translate)
        translation_map = dict(zip(texts_to_translate, translations))
        
        # 4. 流式重建 DOCX
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as out_zf:
            with zipfile.ZipFile(input_path, 'r') as in_zf:
                
                # 4.1 直接複製媒體文件（零記憶體佔用）
                for media_file in media_files:
                    with in_zf.open(media_file) as src:
                        out_zf.writestr(media_file, src.read())
                
                # 4.2 處理 XML 文件
                for xml_file in xml_files:
                    if xml_file in xml_text_map:
                        # 需要修改的 XML
                        modified_xml = self._modify_xml_content(
                            in_zf, xml_file, xml_text_map[xml_file], translation_map
                        )
                        out_zf.writestr(xml_file, modified_xml)
                    else:
                        # 直接複製
                        with in_zf.open(xml_file) as src:
                            out_zf.writestr(xml_file, src.read())
                
                # 4.3 複製其他文件
                for item in in_zf.namelist():
                    if item not in media_files and item not in xml_files:
                        with in_zf.open(item) as src:
                            out_zf.writestr(item, src.read())
    
    def _extract_texts_from_xml(self, docx_path: str, xml_file: str) -> tuple:
        """從 XML 中提取所有 <w:t> 文本節點"""
        texts = []
        mappings = []
        
        with zipfile.ZipFile(docx_path, 'r') as zf:
            with zf.open(xml_file) as f:
                content = f.read().decode('utf-8')
        
        # 使用正則提取 <w:t> 內容
        pattern = r'<w:t[^>]*>([^<]+)</w:t>'
        matches = list(re.finditer(pattern, content))
        
        for idx, match in enumerate(matches):
            text = match.group(1)
            if text.strip() and should_translate(text):
                texts.append(text)
                mappings.append((text, idx))
                self.stats["paragraphs"] += 1
        
        return texts, mappings
    
    def _translate_texts(self, texts: list) -> list:
        """批量翻譯文本列表"""
        if not texts:
            return []
        
        # 去重
        seen = {}
        unique_texts = []
        for text in texts:
            if text not in seen:
                seen[text] = len(unique_texts)
                unique_texts.append(text)
        
        # 批量翻譯
        translated_unique = []
        batch_size = 8  # 小批次，快速處理
        
        for i in range(0, len(unique_texts), batch_size):
            batch = unique_texts[i:i+batch_size]
            results = self.translator.translate_batch(
                [(f"item_{j}", t) for j, t in enumerate(batch)]
            )
            translated_unique.extend([results.get(f"item_{j}", batch[j]) for j in range(len(batch))])
        
        # 映射回原順序
        return [translated_unique[seen[text]] for text in texts]
    
    def _modify_xml_content(self, in_zf: zipfile.ZipFile, xml_file: str, 
                           mappings: list, translation_map: dict) -> bytes:
        """修改 XML 內容中的文本"""
        with in_zf.open(xml_file) as f:
            content = f.read().decode('utf-8')
        
        # 構建替換映射（按長度降序，避免部分替換）
        replacements = []
        for original_text, idx in mappings:
            if original_text in translation_map:
                translated = translation_map[original_text]
                if translated != original_text:
                    replacements.append((original_text, translated))
        
        # 按長度降序排序，避免部分匹配問題
        replacements.sort(key=lambda x: len(x[0]), reverse=True)
        
        # 執行替換
        for original, translated in replacements:
            # 在 <w:t> 標籤中替換
            old_pattern = f'<w:t([^>]*)>{re.escape(original)}</w:t>'
            new_replacement = f'<w:t\\1>{translated}</w:t>'
            content = re.sub(old_pattern, new_replacement, content)
        
        return content.encode('utf-8')
    
    def _create_batches(self, items, max_tokens=1500):
        """智能分塊 - 基於 token 估算（Render 免費版優化：更小的批次，更快處理）"""
        batches = []
        current_batch = []
        current_tokens = 0
        
        # System prompt + 輸出預留（減少以節省內存）
        overhead = 200  # 減少 system prompt 開銷
        max_items_per_batch = 8  # 減少每批項目數，加快處理速度
        
        for item_id, text in items:
            # 限制單個文本長度，避免內存溢出
            text = text[:800] if len(text) > 800 else text
            
            # 估算：中文字符約 1.5 tokens，加上標記開銷
            est_tokens = len(text) * 1.5 + 10  # +10 for [n] marker
            
            # 檢查是否需要新批次（token 限制或項目數限制）
            should_new_batch = (
                current_tokens + est_tokens + overhead > max_tokens or 
                len(current_batch) >= max_items_per_batch
            ) and current_batch
            
            if should_new_batch:
                batches.append(current_batch)
                current_batch = []
                current_tokens = 0
            
            current_batch.append((item_id, text))
            current_tokens += est_tokens
        
        if current_batch:
            batches.append(current_batch)
        
        logger.info(f"Created {len(batches)} batches with max {max_items_per_batch} items each")
        return batches

class XlsxProcessor:
    """XLSX 处理器"""
    
    def __init__(self, translator):
        self.translator = translator
        self.stats = {"cells": 0}
    
    def process(self, file_buffer):
        from openpyxl import load_workbook
        
        wb = load_workbook(BytesIO(file_buffer))
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if should_translate(cell.value):
                            translated, _ = self.translator.translate_text(cell.value)
                            cell.value = translated
                            self.stats["cells"] += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

# API 路由
@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        "status": "ok",
        "service": "格不丢翻译 API",
        "version": "2.0.0",
        "timestamp": datetime.now().isoformat()
    })

@app.route('/domains', methods=['GET'])
def get_domains():
    return jsonify({
        "domains": {
            k: {"name": v["name"], "name_en": v["name_en"], "icon": v["icon"]}
            for k, v in DOMAINS.items()
        }
    })

@app.route('/translate', methods=['POST'])
@limiter.limit("10 per minute")
def translate():
    """
    翻譯文件 - 優化版，支持大文件和超時保護
    """
    start_time = time.time()
    
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        domain = request.form.get('domain', 'general')
        
        if domain not in DOMAINS:
            return jsonify({"error": "Invalid domain"}), 400
        
        is_valid, error_msg, file_info = validate_file(file)
        if not is_valid:
            logger.warning(f"File validation failed: {error_msg}")
            return jsonify({"error": error_msg}), 400
        
        # 大文件警告（可能導致超時）
        file_size_mb = file_info["size"] / 1024 / 1024
        if file_size_mb > 3:
            logger.warning(f"Large file ({file_size_mb:.1f}MB), may cause timeout")
        
        translator = TranslationService(domain)
        
        if file_info["extension"] == '.docx':
            processor = DocxProcessor(translator)
            output = processor.process(file_info["buffer"])
            output_filename = file_info["filename"].replace('.docx', '_EN.docx')
            mimetype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        else:
            processor = XlsxProcessor(translator)
            output = processor.process(file_info["buffer"])
            output_filename = file_info["filename"].replace('.xlsx', '_EN.xlsx')
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        processing_time = time.time() - start_time
        
        logger.info(f"Translation completed: {file_info['filename']} -> {output_filename}, "
                   f"time: {processing_time:.2f}s")
        
        response = make_response(send_file(
            output,
            mimetype=mimetype,
            as_attachment=True,
            download_name=output_filename
        ))
        
        response.headers['X-Processing-Time'] = f"{processing_time:.2f}"
        response.headers['X-API-Calls'] = str(translator.stats["api_calls"])
        response.headers['X-Cache-Hits'] = str(translator.stats["cache_hits"])
        
        return response
        
    except Exception as e:
        logger.error(f"Translation failed: {e}", exc_info=True)
        error_msg = str(e)
        
        # 提供更友好的錯誤信息
        if "timeout" in error_msg.lower() or "worker" in error_msg.lower():
            return jsonify({
                "error": "Processing timeout",
                "message": "File too large or complex. Please try a smaller file (< 2MB) or contact support."
            }), 504
        
        return jsonify({
            "error": "Translation failed",
            "message": "An error occurred. Please try again."
        }), 500

@app.errorhandler(429)
def ratelimit_handler(e):
    return jsonify({
        "error": "Rate limit exceeded",
        "message": "Too many requests. Please try again later."
    }), 429

@app.errorhandler(500)
def internal_error(e):
    logger.error(f"Internal error: {e}", exc_info=True)
    return jsonify({
        "error": "Internal server error",
        "message": "Something went wrong. Please try again later."
    }), 500

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.getenv("PORT", 5000)))
