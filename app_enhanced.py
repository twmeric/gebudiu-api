#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
格不丢翻译 API - 增強版 (v3.0)
集成 Translation Memory + 智能領域檢測

新特性:
- Translation Memory (FAISS + SQLite)
- 自動領域檢測
- 翻譯質量評分
- API成本降低 50-70%
"""

import os
import sys
import logging
from io import BytesIO
from datetime import datetime

from flask import Flask, request, send_file, jsonify, make_response
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from dotenv import load_dotenv

# 導入增強組件
try:
    from enhanced_translation_service import EnhancedTranslationService, TranslationResult
    from translation_memory import TranslationMemory
    from domain_detector import DomainDetector
    from terminology_manager import TerminologyManager, get_terminology_manager
    from terminology_api import init_terminology_routes
    # 格式自學習引擎 (GBD USP)
    from format_fingerprint import get_format_learning_engine, FormatParams
    from layout_analyzer import DocxLayoutAnalyzer, FormatLearningPipeline
    FORMAT_LEARNING_AVAILABLE = True
except ImportError as e:
    FORMAT_LEARNING_AVAILABLE = False
    logging.warning(f"Format learning not available: {e}")

try:
    from enhanced_translation_service import EnhancedTranslationService, TranslationResult
    from translation_memory import TranslationMemory
    from domain_detector import DomainDetector
    from terminology_manager import TerminologyManager, get_terminology_manager
    from terminology_api import init_terminology_routes
    ENHANCED_MODE = True
except ImportError as e:
    ENHANCED_MODE = False
    logging.warning(f"Enhanced modules not available: {e}")
    # 回退到舊版
    from app_secure import app as old_app

# 加载环境变量
load_dotenv()

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('translation_enhanced.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# 創建應用
app = Flask(__name__, static_folder='.', static_url_path='')

# CORS配置
CORS(app, origins="*", supports_credentials=False)

@app.after_request
def after_request(response):
    """為所有響應添加 CORS 頭"""
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization,X-Requested-With')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# 速率限制
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# 配置文件
MAX_FILE_SIZE = 16 * 1024 * 1024
ALLOWED_EXTENSIONS = {'.docx', '.xlsx'}

# 领域配置
DOMAINS = {
    "general": {"name": "通用领域", "name_en": "General", "icon": "📋"},
    "electronics": {"name": "电子产品", "name_en": "Electronics", "icon": "🔌"},
    "medical": {"name": "医疗器械", "name_en": "Medical", "icon": "🏥"},
    "legal": {"name": "法律合同", "name_en": "Legal", "icon": "⚖️"},
    "marketing": {"name": "市场营销", "name_en": "Marketing", "icon": "📢"},
    "industrial": {"name": "工业技术", "name_en": "Industrial", "icon": "🏭"},
    "software": {"name": "软件/IT", "name_en": "Software", "icon": "💻"}
}

# 全局服務實例
translation_service = None
translation_memory = None
terminology_manager = None

def init_services():
    """初始化翻譯服務"""
    global translation_service, translation_memory, terminology_manager
    
    if not ENHANCED_MODE:
        logger.warning("Running in legacy mode (enhanced features disabled)")
        return
    
    try:
        # 初始化Translation Memory
        translation_memory = TranslationMemory()
        logger.info("Translation Memory initialized")
        
        # 初始化術語表管理器
        terminology_manager = get_terminology_manager()
        logger.info("Terminology Manager initialized")
        
        # 初始化增強翻譯服務
        translation_service = EnhancedTranslationService(
            domain="general",
            use_tm=True,
            tm_threshold=0.85,
            auto_detect_domain=True,
            use_terminology=True
        )
        logger.info("Enhanced Translation Service initialized")
        
        # 初始化術語表API路由
        init_terminology_routes(app, terminology_manager)
        
        # 初始化格式自學習引擎 (GBD USP)
        if FORMAT_LEARNING_AVAILABLE:
            format_engine = get_format_learning_engine()
            logger.info("Format Learning Engine initialized (GBD USP)")
        
    except Exception as e:
        logger.error(f"Failed to initialize enhanced services: {e}")
        logger.info("Falling back to legacy mode")

def validate_file(file_storage):
    """验证上传文件"""
    if not file_storage or not getattr(file_storage, 'filename', None):
        return False, "No file selected", None
    
    filename = os.path.basename(file_storage.filename)
    ext = os.path.splitext(filename)[1].lower()
    
    # 詳細日誌用於調試
    logger.info(f"Validating file: {filename}, extension: {ext}")
    
    # 支持更多變體（微信/瀏覽器可能修改文件名）
    valid_extensions = {'.docx', '.xlsx', '.doc', '.xls'}
    
    if ext not in valid_extensions:
        logger.warning(f"Invalid file extension: {ext} for file {filename}")
        return False, f"Invalid file type: {ext}. Allowed: .docx, .xlsx", None
    
    # 如果是舊格式 .doc 或 .xls，給出明確提示
    if ext in ['.doc', '.xls']:
        return False, f"舊格式 {ext} 不支持，請轉換為 {ext}x 格式後重試", None
    
    try:
        file_buffer = file_storage.read()
        file_storage.seek(0)
    except Exception as e:
        logger.error(f"Failed to read file: {e}")
        return False, "Failed to read uploaded file", None
    
    if len(file_buffer) == 0:
        return False, "File is empty", None
    
    if len(file_buffer) > MAX_FILE_SIZE:
        max_mb = MAX_FILE_SIZE / 1024 / 1024
        return False, f"File too large. Maximum size: {max_mb}MB", None
    
    # 文件格式檢查
    header = file_buffer[:4]
    
    # 檢測是否為舊格式 .doc (OLE Compound File)
    old_doc_header = b'\xd0\xcf\x11\xe0'  # .doc 文件頭
    if header == old_doc_header:
        logger.warning(f"Old .doc format detected: {header.hex()}")
        return False, "檢測到舊格式 .doc 文件，請將文件另存為 .docx 格式後重試", None
    
    # ZIP格式檢查（.docx 和 .xlsx 都是 ZIP）
    valid_zip_headers = [b'PK\x03\x04', b'PK\x05\x06', b'PK\x07\x08']
    if header not in valid_zip_headers:
        logger.warning(f"Invalid file header: {header.hex()}")
        return False, f"文件格式錯誤（文件頭: {header.hex()}），請上傳有效的 .docx 或 .xlsx 文件", None
    
    return True, None, {
        "filename": filename,
        "extension": ext,
        "size": len(file_buffer),
        "buffer": file_buffer
    }

def should_translate(text):
    """判斷文本是否需要翻譯"""
    if not text or len(text.strip()) < 2:
        return False
    text = text.strip()
    if text.isdigit():
        return False
    if text.startswith(('http://', 'https://')):
        return False
    return True

# API 路由
@app.route('/')
def index():
    """用戶指南主頁 - 提供 API 文檔"""
    try:
        return app.send_static_file('index.html')
    except:
        # 如果靜態文件不存在，返回基本信息
        return jsonify({
            "service": "GeBuDiu API - 专业 DOCX 文档翻译",
            "version": "3.0.0",
            "tagline": "越翻译，格式越精准",
            "documentation": "/docs",
            "health": "/health"
        })

@app.route('/docs')
def docs():
    """用戶指南頁面"""
    try:
        return app.send_static_file('index.html')
    except:
        return jsonify({"error": "Documentation not available"}), 503

@app.route('/health', methods=['GET'])
def health():
    """健康檢查"""
    status = {
        "status": "ok",
        "service": "格不丢翻译 API - Enhanced",
        "version": "3.0.0",
        "enhanced_mode": ENHANCED_MODE,
        "timestamp": datetime.now().isoformat()
    }
    
    if ENHANCED_MODE and translation_service:
        status["features"] = {
            "translation_memory": True,
            "domain_detection": True,
            "quality_scoring": True,
            "terminology_management": True,
            "format_learning": FORMAT_LEARNING_AVAILABLE  # GBD USP
        }
        status["stats"] = translation_service.get_stats_report()
        
        # 添加格式學習狀態
        if FORMAT_LEARNING_AVAILABLE:
            try:
                format_engine = get_format_learning_engine()
                format_stats = format_engine.get_learning_stats()
                status["format_learning"] = {
                    "status": "active",
                    "tagline": "每一翻譯，都讓下一個更完美",
                    "patterns_learned": format_stats.get("total_patterns", 0),
                    "optimizations_made": format_stats.get("total_optimizations", 0)
                }
            except:
                status["format_learning"] = {"status": "initializing"}
    
    return jsonify(status)

@app.route('/domains', methods=['GET'])
def get_domains():
    """獲取支持的領域列表"""
    return jsonify({
        "domains": {
            k: {"name": v["name"], "name_en": v["name_en"], "icon": v["icon"]}
            for k, v in DOMAINS.items()
        }
    })

@app.route('/translate', methods=['POST'])
@limiter.limit("10 per minute")
def translate():
    """
    翻譯文件 - 增強版
    
    Form參數:
    - file: 要上傳的文件
    - domain: 領域 (可選, 默認自動檢測)
    - format: 輸出格式 (保留)
    """
    import time
    start_time = time.time()
    
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        domain = request.form.get('domain', 'general')
        
        if domain not in DOMAINS:
            return jsonify({"error": "Invalid domain"}), 400
        
        is_valid, error_msg, file_info = validate_file(file)
        if not is_valid:
            logger.warning(f"File validation failed: {error_msg}")
            return jsonify({"error": error_msg}), 400
        
        file_size_mb = file_info["size"] / 1024 / 1024
        if file_size_mb > 3:
            logger.warning(f"Large file ({file_size_mb:.1f}MB), processing may take time")
        
        # 使用增強服務或回退到舊版
        if ENHANCED_MODE and translation_service:
            output, stats = process_with_enhanced_service(file_info, domain)
        else:
            output, stats = process_with_legacy_service(file_info, domain)
        
        processing_time = time.time() - start_time
        
        logger.info(f"Translation completed: {file_info['filename']}, time: {processing_time:.2f}s")
        
        # 構建響應
        response = make_response(send_file(
            output,
            mimetype=get_mimetype(file_info["extension"]),
            as_attachment=True,
            download_name=file_info["filename"].replace(file_info["extension"], f'_EN{file_info["extension"]}')
        ))
        
        response.headers['X-Processing-Time'] = f"{processing_time:.2f}"
        if stats:
            response.headers['X-API-Calls'] = str(stats.get("api_calls", 0))
            response.headers['X-TM-Hits'] = str(stats.get("tm_hits", 0))
            response.headers['X-Cache-Hits'] = str(stats.get("cache_hits", 0))
            response.headers['X-Domain'] = stats.get("domain", domain)
        
        return response
        
    except Exception as e:
        logger.error(f"Translation failed: {e}", exc_info=True)
        return jsonify({
            "error": "Translation failed",
            "message": str(e)
        }), 500

def process_with_enhanced_service(file_info, domain):
    """使用增強服務處理文件"""
    from enhanced_docx_processor import EnhancedDocxProcessor
    
    # 更新服務領域
    translation_service.domain = domain
    
    if file_info["extension"] == '.docx':
        processor = EnhancedDocxProcessor(translation_service)
        output = processor.process(file_info["buffer"], file_info["filename"])
    else:
        # XLSX處理
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_info["buffer"]))
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if should_translate(cell.value):
                            translated, is_tm = translation_service.translate_text(cell.value)
                            cell.value = translated
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
    
    stats = translation_service.get_stats_report()
    return output, stats

def process_with_legacy_service(file_info, domain):
    """回退到舊版服務"""
    # 這裡可以調用舊版處理邏輯
    # 暫時拋出錯誤，提示使用新版
    raise Exception("Enhanced mode required. Please check dependencies.")

def get_mimetype(extension):
    """獲取MIME類型"""
    if extension == '.docx':
        return 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    else:
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

@app.route('/tm/stats', methods=['GET'])
def tm_stats():
    """Translation Memory 統計"""
    if not ENHANCED_MODE or not translation_memory:
        return jsonify({"error": "Translation Memory not available"}), 503
    
    try:
        stats = translation_memory.get_stats()
        return jsonify(stats)
    except Exception as e:
        logger.error(f"Failed to get TM stats: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/tm/search', methods=['POST'])
def tm_search():
    """搜索Translation Memory"""
    if not ENHANCED_MODE or not translation_memory:
        return jsonify({"error": "Translation Memory not available"}), 503
    
    try:
        data = request.get_json()
        query = data.get('query', '')
        domain = data.get('domain')
        
        if not query:
            return jsonify({"error": "Query is required"}), 400
        
        results = translation_memory.search(query, domain)
        
        return jsonify({
            "query": query,
            "results": [
                {
                    "source": r.source,
                    "target": r.target,
                    "similarity": r.similarity,
                    "domain": r.domain,
                    "hit_count": r.hit_count
                }
                for r in results
            ]
        })
        
    except Exception as e:
        logger.error(f"TM search failed: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/detect-domain', methods=['POST'])
def detect_domain():
    """檢測文檔領域"""
    if not ENHANCED_MODE:
        return jsonify({"error": "Domain detection not available"}), 503
    
    try:
        data = request.get_json()
        filename = data.get('filename', '')
        samples = data.get('samples', [])
        
        detector = DomainDetector()
        domain, confidence = detector.detect(filename, samples)
        info = detector.get_domain_info(domain)
        
        return jsonify({
            "domain": domain,
            "confidence": confidence,
            "info": info
        })
        
    except Exception as e:
        logger.error(f"Domain detection failed: {e}")
        return jsonify({"error": str(e)}), 500


# ==================== 格式自學習引擎 API (GBD USP) ====================

@app.route('/format-learning/stats', methods=['GET'])
def format_learning_stats():
    """獲取格式學習統計 - 展示 GBD 的 USP"""
    if not FORMAT_LEARNING_AVAILABLE:
        return jsonify({
            "status": "not_available",
            "message": "Format learning engine not available"
        }), 503
    
    try:
        engine = get_format_learning_engine()
        stats = engine.get_learning_stats()
        return jsonify(stats)
    except Exception as e:
        logger.error(f"Failed to get format learning stats: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/format-learning/analyze', methods=['POST'])
def format_learning_analyze():
    """
    分析原文和譯文的佈局差異
    這是「越翻譯越聰明」的核心功能
    """
    if not FORMAT_LEARNING_AVAILABLE:
        return jsonify({"error": "Format learning not available"}), 503
    
    try:
        if 'source' not in request.files or 'translated' not in request.files:
            return jsonify({"error": "需要上傳原文(source)和譯文(translated)文件"}), 400
        
        source_file = request.files['source']
        translated_file = request.files['translated']
        domain = request.form.get('domain', 'general')
        
        # 讀取文件
        source_bytes = source_file.read()
        translated_bytes = translated_file.read()
        
        if len(source_bytes) == 0 or len(translated_bytes) == 0:
            return jsonify({"error": "文件為空"}), 400
        
        # 獲取當前使用的格式參數（從翻譯服務或默認值）
        params_used = FormatParams(
            font_size=float(request.form.get('font_size', 11.0)),
            line_spacing=float(request.form.get('line_spacing', 1.15)),
            paragraph_spacing=float(request.form.get('paragraph_spacing', 6.0)),
            margin_cm=float(request.form.get('margin_cm', 2.5))
        )
        
        # 執行分析流水線
        engine = get_format_learning_engine()
        pipeline = FormatLearningPipeline(engine)
        
        result = pipeline.process_translation(
            source_bytes, translated_bytes, domain, params_used
        )
        
        return jsonify({
            "status": "success",
            "message": "分析完成，已記錄到學習引擎",
            "result": result
        })
        
    except Exception as e:
        logger.error(f"Format learning analysis failed: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/format-learning/predict', methods=['POST'])
def format_learning_predict():
    """
    預測最佳格式參數
    基於歷史相似內容的最佳表現
    """
    if not FORMAT_LEARNING_AVAILABLE:
        return jsonify({"error": "Format learning not available"}), 503
    
    try:
        data = request.get_json()
        
        # 構建內容指紋
        from format_fingerprint import ContentFingerprint
        fingerprint = ContentFingerprint(
            domain=data.get('domain', 'general'),
            total_chars=data.get('total_chars', 1000),
            avg_sentence_length=data.get('avg_sentence_length', 50),
            paragraph_count=data.get('paragraph_count', 10),
            table_count=data.get('table_count', 0),
            image_count=data.get('image_count', 0),
            structure_complexity=data.get('structure_complexity', 0.3)
        )
        
        engine = get_format_learning_engine()
        optimal_params = engine.predict_optimal_params(fingerprint)
        
        return jsonify({
            "status": "success",
            "predicted_params": optimal_params.to_dict(),
            "fingerprint": fingerprint.to_dict()
        })
        
    except Exception as e:
        logger.error(f"Format prediction failed: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/format-learning/report', methods=['GET'])
def format_learning_report():
    """
    生成格式學習報告
    展示「越翻譯越聰明」的累積效果
    """
    if not FORMAT_LEARNING_AVAILABLE:
        return jsonify({"error": "Format learning not available"}), 503
    
    try:
        engine = get_format_learning_engine()
        stats = engine.get_learning_stats()
        
        # 生成用戶友好的報告
        report = {
            "title": "GeBuDiu 格式智能學習報告",
            "tagline": "每一翻譯，都讓下一個更完美",
            "summary": {
                "total_documents_analyzed": stats.get("total_optimizations", 0),
                "learning_patterns_stored": stats.get("total_patterns", 0),
                "average_satisfaction": f"{stats.get('average_satisfaction', 0) * 100:.0f}%",
                "status": "學習中" if stats.get("total_patterns", 0) < 10 else "已成熟"
            },
            "insights": [],
            "domain_performance": stats.get("domain_breakdown", [])
        }
        
        # 生成洞察
        if stats.get("total_patterns", 0) > 0:
            report["insights"].append(f"已為您累積學習 {stats['total_patterns']} 種文件格式模式")
            report["insights"].append(f"當前平均格式滿意度: {report['summary']['average_satisfaction']}")
        
        if stats.get("domain_breakdown"):
            best_domain = max(stats["domain_breakdown"], key=lambda x: x.get("avg_score", 0))
            report["insights"].append(f"最擅長處理: {best_domain.get('domain', 'general')} 領域文件")
        
        if stats.get("total_patterns", 0) < 5:
            report["insights"].append("💡 建議: 多翻譯幾份文件以建立個人化的格式偏好")
        
        return jsonify(report)
        
    except Exception as e:
        logger.error(f"Format learning report failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.errorhandler(429)
def ratelimit_handler(e):
    return jsonify({
        "error": "Rate limit exceeded",
        "message": "Too many requests. Please try again later."
    }), 429

@app.errorhandler(500)
def internal_error(e):
    logger.error(f"Internal error: {e}", exc_info=True)
    return jsonify({
        "error": "Internal server error",
        "message": "Something went wrong. Please try again later."
    }), 500

# 初始化
init_services()

if __name__ == '__main__':
    port = int(os.getenv("PORT", 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
